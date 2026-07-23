# inventario/views.py
# VERSIÓN COMPLETA — reemplaza el archivo existente

from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render, get_object_or_404
from django.db.models import F, Q, Sum, DecimalField, Prefetch
from django.utils import timezone
from django.core.exceptions import ValidationError
from decimal import Decimal

from rest_framework import viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser

from .models import (
    Articulo, Marca, Rubro, Deposito,
    BalanceStock, MovimientoStockLedger,
    AjusteStock, ItemAjusteStock,
    TransferenciaInterna, ItemTransferencia,
    MotivoAjuste, TipoStock, ProveedorArticulo,
)
from .serializers import (
    # Artículo
    ArticuloListSerializer,
    ArticuloDetailSerializer,
    ArticuloWriteSerializer,
    # Auxiliares
    MarcaSerializer,
    RubroSerializer,
    CategoriaImpositivaSerializer,
    DepositoSerializer,
    MotivoAjusteSerializer,
    TipoStockSerializer,
    # Ajustes
    AjusteStockSerializer,
    AjusteStockListSerializer,
    # Transferencias
    TransferenciaSerializer,
    TransferenciaListSerializer,
    # Ledger
    LedgerSerializer,
    # Balance
    BalanceStockSerializer,
    # Proveedor de artículo
    ProveedorArticuloSerializer,
)
from .filters import ArticuloSearchFilter
from .services import AjusteService, TransferenciaService

from parametros.models import CategoriaImpositiva


# ─────────────────────────────────────────────
#  ARTÍCULOS
# ─────────────────────────────────────────────

class ArticuloViewSet(viewsets.ModelViewSet):
    """
    ViewSet completo para Artículos.
    ...
    """

    # Soporta tanto JSON (sin foto) como multipart/form-data (con foto)
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    filter_backends = [ArticuloSearchFilter]

    def get_queryset(self):
        qs = Articulo.objects.select_related(
            'marca', 'rubro', 'categoria_impositiva',
            'precio_costo_moneda', 'precio_venta_moneda',
            'unidad_medida_stock', 'unidad_medida_venta',
        ).prefetch_related(
            'impuestos',
            'stocks__deposito',
            'balances_stock__tipo_stock',
            'balances_stock__deposito',
        ).order_by('cod_articulo')

        # Filtros adicionales por querystring
        rubro_id = self.request.query_params.get('rubro')
        marca_id = self.request.query_params.get('marca')
        activo = self.request.query_params.get('activo')
        perfil = self.request.query_params.get('perfil')
        administra_stock = self.request.query_params.get('administra_stock')
        bajo_minimo = self.request.query_params.get('bajo_minimo')

        if rubro_id:
            qs = qs.filter(rubro_id=rubro_id)
        if marca_id:
            qs = qs.filter(marca_id=marca_id)
        if activo is not None and activo != '':
            qs = qs.filter(is_active=(activo.lower() == 'true'))
        if perfil:
            qs = qs.filter(perfil=perfil)
        if administra_stock is not None and administra_stock != '':
            qs = qs.filter(administra_stock=(administra_stock.lower() == 'true'))

        # Filtro por proveedor: devuelve solo artículos vinculados a ese proveedor
        proveedor_id = self.request.query_params.get('proveedor')
        if proveedor_id:
            qs = qs.filter(proveedores__pk=proveedor_id).distinct()

        # bajo_minimo requiere proceso en Python (propiedad calculada)
        return qs

    def get_serializer_class(self):
        if self.action == 'list':
            return ArticuloListSerializer
        if self.action in ['create', 'update', 'partial_update']:
            return ArticuloWriteSerializer
        return ArticuloDetailSerializer

    def destroy(self, request, *args, **kwargs):
        """
        Nunca eliminar físicamente un artículo.
        Solo desactivación lógica.
        """
        articulo = self.get_object()
        articulo.is_active = False
        articulo.save(update_fields=['is_active'])
        return Response(
            {'detail': f"Artículo '{articulo.descripcion}' desactivado correctamente."},
            status=status.HTTP_200_OK
        )

    @action(detail=True, methods=['post'], url_path='desactivar')
    def desactivar(self, request, pk=None):
        articulo = self.get_object()
        articulo.is_active = False
        articulo.save(update_fields=['is_active'])
        return Response({'detail': 'Artículo desactivado.'}, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='activar')
    def activar(self, request, pk=None):
        articulo = self.get_object()
        articulo.is_active = True
        articulo.save(update_fields=['is_active'])
        return Response({'detail': 'Artículo activado.'}, status=status.HTTP_200_OK)

    @action(detail=True, methods=['get'], url_path='stock')
    def stock(self, request, pk=None):
        """
        Devuelve el balance de stock del artículo por depósito y tipo.
        """
        articulo = self.get_object()
        balances = BalanceStock.objects.filter(
            articulo=articulo
        ).select_related('deposito', 'tipo_stock').order_by('deposito__nombre', 'tipo_stock__codigo')

        serializer = BalanceStockSerializer(balances, many=True)
        return Response({
            'articulo_id': articulo.pk,
            'cod_articulo': articulo.cod_articulo,
            'descripcion': articulo.descripcion,
            'stock_total': articulo.stock_total,
            'stock_disponible': articulo.stock_disponible_calculado,
            'necesita_reposicion': articulo.necesita_reposicion,
            'stock_minimo': articulo.stock_minimo,
            'balances': serializer.data,
        })

    @action(detail=True, methods=['get'], url_path='kardex')
    def kardex(self, request, pk=None):
        """
        Kardex del artículo: historial de movimientos con saldo acumulado.
        """
        articulo = self.get_object()
        movimientos = MovimientoStockLedger.objects.filter(
            articulo=articulo
        ).select_related('deposito', 'tipo_stock', 'usuario')

        # Filtros opcionales
        deposito_id = request.query_params.get('deposito')
        desde = request.query_params.get('desde')
        hasta = request.query_params.get('hasta')

        if deposito_id:
            movimientos = movimientos.filter(deposito_id=deposito_id)
        if desde:
            movimientos = movimientos.filter(fecha_movimiento__date__gte=desde)
        if hasta:
            movimientos = movimientos.filter(fecha_movimiento__date__lte=hasta)

        movimientos = movimientos.order_by('fecha_movimiento', 'pk')

        # Calcular saldo acumulado (solo stock REAL/físico)
        filas = []
        saldo = 0
        for mov in movimientos:
            impacto = 0
            if mov.tipo_stock and mov.tipo_stock.es_fisico:
                impacto = float(mov.cantidad)
            saldo += impacto

            filas.append({
                'id': mov.pk,
                'fecha': mov.fecha_movimiento,
                'deposito': mov.deposito.nombre if mov.deposito else '',
                'tipo_codigo': mov.tipo_stock.codigo if mov.tipo_stock else '',
                'tipo_nombre': mov.tipo_stock.nombre if mov.tipo_stock else '',
                'origen_sistema': mov.origen_sistema,
                'origen_referencia': mov.origen_referencia,
                'entrada': float(mov.cantidad) if mov.cantidad > 0 else 0,
                'salida': float(abs(mov.cantidad)) if mov.cantidad < 0 else 0,
                'saldo': round(saldo, 3),
                'usuario': (
                    mov.usuario.get_full_name() or mov.usuario.username
                ) if mov.usuario else None,
                'observaciones': mov.observaciones,
            })

        return Response({
            'articulo_id': articulo.pk,
            'cod_articulo': articulo.cod_articulo,
            'descripcion': articulo.descripcion,
            'saldo_final': round(saldo, 3),
            'total_movimientos': len(filas),
            'kardex': filas,
        })

    @action(detail=False, methods=['get'], url_path='choices')
    def choices(self, request):
        return Response({
            'perfil': Articulo.Perfil.choices,
        })

    @action(detail=False, methods=['get'], url_path='alertas')
    def alertas(self, request):
        """
        Devuelve artículos que necesitan reposición (stock <= stock_minimo).
        """
        articulos = self.get_queryset().filter(
            administra_stock=True,
            is_active=True,
            stock_minimo__gt=0,
        )
        # Filtrar con la propiedad calculada en Python
        alertas = [a for a in articulos if a.necesita_reposicion]
        serializer = ArticuloListSerializer(alertas, many=True)
        return Response({
            'count': len(alertas),
            'results': serializer.data,
        })

    @action(detail=False, methods=['get'], url_path='dashboard')
    def dashboard(self, request):
        """
        KPIs de inventario para el panel principal.
        Responde todas las métricas en una sola llamada.
        """
        total_articulos = Articulo.objects.filter(is_active=True).count()
        sin_stock = 0
        bajo_minimo = 0
        valor_total = 0

        articulos_con_stock = Articulo.objects.filter(
            is_active=True, administra_stock=True
        ).prefetch_related('stocks', 'balances_stock__tipo_stock')

        for a in articulos_con_stock:
            disp = float(a.stock_disponible_calculado)
            if disp <= 0:
                sin_stock += 1
            if a.necesita_reposicion:
                bajo_minimo += 1
            valor_total += disp * float(a.precio_costo_monto or 0)

        ultimos_movimientos = MovimientoStockLedger.objects.select_related(
            'articulo', 'deposito', 'tipo_stock'
        ).order_by('-fecha_registro')[:8]

        movimientos_data = LedgerSerializer(ultimos_movimientos, many=True).data

        return Response({
            'total_articulos_activos': total_articulos,
            'articulos_sin_stock': sin_stock,
            'articulos_bajo_minimo': bajo_minimo,
            'valor_stock_total': round(valor_total, 2),
            'ultimos_movimientos': movimientos_data,
        })




# ─────────────────────────────────────────────
#  PROVEEDOR DE ARTÍCULO (CRUD nested)
# ─────────────────────────────────────────────

class ProveedorArticuloViewSet(viewsets.ModelViewSet):
    """
    CRUD de relaciones Artículo ↔ Proveedor, nested bajo /api/articulos/{articulo_pk}/proveedores/

    GET    /api/articulos/{articulo_pk}/proveedores/         → lista de proveedores del artículo
    POST   /api/articulos/{articulo_pk}/proveedores/         → agregar proveedor
    GET    /api/articulos/{articulo_pk}/proveedores/{id}/    → detalle
    PATCH  /api/articulos/{articulo_pk}/proveedores/{id}/    → editar cod/desc/fuente_de_verdad
    DELETE /api/articulos/{articulo_pk}/proveedores/{id}/    → quitar proveedor del artículo

    POST   /api/articulos/{articulo_pk}/proveedores/{id}/set_fuente_de_verdad/
           → marcar como fuente de verdad de precio de costo
    """
    serializer_class = ProveedorArticuloSerializer

    def get_queryset(self):
        articulo_pk = self.kwargs.get('articulo_pk')
        return ProveedorArticulo.objects.filter(
            articulo_id=articulo_pk
        ).select_related('proveedor__entidad').order_by(
            '-es_fuente_de_verdad', 'proveedor__entidad__razon_social'
        )

    def perform_create(self, serializer):
        articulo_pk = self.kwargs.get('articulo_pk')
        articulo = get_object_or_404(Articulo, pk=articulo_pk)
        serializer.save(articulo=articulo)

    @action(detail=True, methods=['post'], url_path='set_fuente_de_verdad')
    def set_fuente_de_verdad(self, request, articulo_pk=None, pk=None):
        """
        Marca este proveedor como la fuente de verdad del precio de costo.
        Automáticamente desmarca al proveedor anterior (lógica en el model.save()).
        """
        pa = self.get_object()
        pa.es_fuente_de_verdad = True
        pa.save()
        return Response(
            {'detail': f'{pa.proveedor.entidad.razon_social} marcado como fuente de verdad de precio.'},
            status=status.HTTP_200_OK
        )

# ─────────────────────────────────────────────
#  MAESTROS
# ─────────────────────────────────────────────

class MarcaViewSet(viewsets.ModelViewSet):
    queryset = Marca.objects.all().order_by('nombre')
    serializer_class = MarcaSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['nombre']


class RubroViewSet(viewsets.ModelViewSet):
    queryset = Rubro.objects.all().order_by('nombre')
    serializer_class = RubroSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['nombre']


class CategoriaImpositivaViewSet(viewsets.ModelViewSet):
    queryset = CategoriaImpositiva.objects.all()
    serializer_class = CategoriaImpositivaSerializer


class DepositoViewSet(viewsets.ModelViewSet):
    queryset = Deposito.objects.all().order_by('nombre')
    serializer_class = DepositoSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['nombre']


class MotivoAjusteViewSet(viewsets.ModelViewSet):
    queryset = MotivoAjuste.objects.all().order_by('nombre')
    serializer_class = MotivoAjusteSerializer


class TipoStockViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = TipoStock.objects.all().order_by('codigo')
    serializer_class = TipoStockSerializer


# ─────────────────────────────────────────────
#  AJUSTES DE STOCK
# ─────────────────────────────────────────────

class AjusteStockViewSet(viewsets.ModelViewSet):
    """
    CRUD de ajustes manuales de stock.

    GET    /api/inventario/ajustes/           → lista
    POST   /api/inventario/ajustes/           → crear (en Borrador)
    GET    /api/inventario/ajustes/{id}/      → detalle
    PATCH  /api/inventario/ajustes/{id}/      → editar (solo si es Borrador)
    DELETE /api/inventario/ajustes/{id}/      → eliminar (solo si es Borrador)

    Acciones:
    POST /api/inventario/ajustes/{id}/confirmar/ → aplica stock
    POST /api/inventario/ajustes/{id}/anular/    → revierte si estaba confirmado
    """

    filter_backends = [filters.OrderingFilter]
    ordering = ['-fecha']

    def get_queryset(self):
        qs = AjusteStock.objects.select_related(
            'deposito', 'motivo', 'created_by'
        ).prefetch_related('items__articulo').order_by('-fecha')

        deposito_id = self.request.query_params.get('deposito')
        estado = self.request.query_params.get('estado')
        motivo_id = self.request.query_params.get('motivo')
        desde = self.request.query_params.get('desde')
        hasta = self.request.query_params.get('hasta')

        if deposito_id:
            qs = qs.filter(deposito_id=deposito_id)
        if estado:
            qs = qs.filter(estado=estado)
        if motivo_id:
            qs = qs.filter(motivo_id=motivo_id)
        if desde:
            qs = qs.filter(fecha__date__gte=desde)
        if hasta:
            qs = qs.filter(fecha__date__lte=hasta)

        return qs

    def get_serializer_class(self):
        if self.action == 'list':
            return AjusteStockListSerializer
        return AjusteStockSerializer

    def update(self, request, *args, **kwargs):
        ajuste = self.get_object()
        if ajuste.estado != 'BR':
            return Response(
                {'detail': 'Solo se pueden editar ajustes en estado Borrador.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        ajuste = self.get_object()
        if ajuste.estado != 'BR':
            return Response(
                {'detail': 'Solo se pueden eliminar ajustes en estado Borrador.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['post'], url_path='confirmar')
    def confirmar(self, request, pk=None):
        ajuste = self.get_object()
        try:
            AjusteService.confirmar_ajuste(ajuste)
            return Response(
                {'detail': f'Ajuste #{ajuste.pk} confirmado. Stock actualizado.'},
                status=status.HTTP_200_OK
            )
        except ValidationError as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'], url_path='anular')
    def anular(self, request, pk=None):
        ajuste = self.get_object()
        if ajuste.estado == 'AN':
            return Response(
                {'detail': 'El ajuste ya está anulado.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if ajuste.estado == 'BR':
            # Borrador: solo cambiar estado, no hay stock que revertir
            ajuste.estado = 'AN'
            ajuste.save(update_fields=['estado'])
            return Response({'detail': 'Ajuste anulado.'}, status=status.HTTP_200_OK)
        # Confirmado: hay que revertir el stock
        # Por ahora marcamos como anulado; la lógica de reversión se puede
        # agregar en AjusteService si se necesita.
        ajuste.estado = 'AN'
        ajuste.save(update_fields=['estado'])
        return Response(
            {'detail': 'Ajuste anulado. Nota: la reversión de stock debe hacerse con un nuevo ajuste compensatorio.'},
            status=status.HTTP_200_OK
        )


# ─────────────────────────────────────────────
#  TRANSFERENCIAS ENTRE DEPÓSITOS
# ─────────────────────────────────────────────

class TransferenciaViewSet(viewsets.ModelViewSet):
    """
    CRUD de transferencias entre depósitos.

    Acciones de flujo:
    POST /api/inventario/transferencias/{id}/enviar/   → despacha mercadería (Borrador → En Tránsito)
    POST /api/inventario/transferencias/{id}/recibir/  → recibe mercadería (En Tránsito → Completada)
    POST /api/inventario/transferencias/{id}/anular/   → anula (solo Borrador)
    """

    def get_queryset(self):
        qs = TransferenciaInterna.objects.select_related(
            'origen', 'destino', 'created_by'
        ).prefetch_related('items__articulo').order_by('-fecha')

        estado = self.request.query_params.get('estado')
        origen_id = self.request.query_params.get('origen')
        destino_id = self.request.query_params.get('destino')
        desde = self.request.query_params.get('desde')
        hasta = self.request.query_params.get('hasta')

        if estado:
            qs = qs.filter(estado=estado)
        if origen_id:
            qs = qs.filter(origen_id=origen_id)
        if destino_id:
            qs = qs.filter(destino_id=destino_id)
        if desde:
            qs = qs.filter(fecha__date__gte=desde)
        if hasta:
            qs = qs.filter(fecha__date__lte=hasta)

        return qs

    def get_serializer_class(self):
        if self.action == 'list':
            return TransferenciaListSerializer
        return TransferenciaSerializer

    def update(self, request, *args, **kwargs):
        transferencia = self.get_object()
        if transferencia.estado != 'BR':
            return Response(
                {'detail': 'Solo se pueden editar transferencias en estado Borrador.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        transferencia = self.get_object()
        if transferencia.estado != 'BR':
            return Response(
                {'detail': 'Solo se pueden eliminar transferencias en estado Borrador.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['post'], url_path='enviar')
    def enviar(self, request, pk=None):
        transferencia = self.get_object()
        try:
            TransferenciaService.despachar_transferencia(transferencia)
            return Response(
                {'detail': f'Transferencia #{transferencia.pk} enviada. Mercadería en tránsito.'},
                status=status.HTTP_200_OK
            )
        except ValidationError as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'], url_path='recibir')
    def recibir(self, request, pk=None):
        transferencia = self.get_object()
        try:
            TransferenciaService.recibir_transferencia(transferencia)
            return Response(
                {'detail': f'Transferencia #{transferencia.pk} recibida. Stock actualizado en depósito destino.'},
                status=status.HTTP_200_OK
            )
        except ValidationError as e:
            return Response({'detail': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'], url_path='anular')
    def anular(self, request, pk=None):
        transferencia = self.get_object()
        if transferencia.estado != 'BR':
            return Response(
                {'detail': 'Solo se pueden anular transferencias en estado Borrador.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        transferencia.estado = 'AN'
        transferencia.save(update_fields=['estado'])
        return Response({'detail': 'Transferencia anulada.'}, status=status.HTTP_200_OK)


# ─────────────────────────────────────────────
#  LEDGER (historial inmutable, solo lectura)
# ─────────────────────────────────────────────

class LedgerViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Consulta del Ledger de stock. Solo lectura — nunca se escribe desde aquí.
    """
    serializer_class = LedgerSerializer

    def get_queryset(self):
        qs = MovimientoStockLedger.objects.select_related(
            'articulo', 'deposito', 'tipo_stock', 'usuario'
        ).order_by('-fecha_registro')

        articulo_id = self.request.query_params.get('articulo')
        deposito_id = self.request.query_params.get('deposito')
        origen = self.request.query_params.get('origen_sistema')
        desde = self.request.query_params.get('desde')
        hasta = self.request.query_params.get('hasta')

        if articulo_id:
            qs = qs.filter(articulo_id=articulo_id)
        if deposito_id:
            qs = qs.filter(deposito_id=deposito_id)
        if origen:
            qs = qs.filter(origen_sistema=origen)
        if desde:
            qs = qs.filter(fecha_movimiento__date__gte=desde)
        if hasta:
            qs = qs.filter(fecha_movimiento__date__lte=hasta)

        return qs


# ─────────────────────────────────────────────
#  BALANCE DE STOCK (solo lectura)
# ─────────────────────────────────────────────

class BalanceStockViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Consulta del balance actual de stock por artículo/depósito/tipo.
    """
    serializer_class = BalanceStockSerializer

    def get_queryset(self):
        qs = BalanceStock.objects.select_related(
            'articulo', 'deposito', 'tipo_stock'
        )

        articulo_id = self.request.query_params.get('articulo')
        deposito_id = self.request.query_params.get('deposito')
        solo_fisico = self.request.query_params.get('solo_fisico')
        solo_positivos = self.request.query_params.get('solo_positivos')

        if articulo_id:
            qs = qs.filter(articulo_id=articulo_id)
        if deposito_id:
            qs = qs.filter(deposito_id=deposito_id)
        if solo_fisico and solo_fisico.lower() == 'true':
            qs = qs.filter(tipo_stock__es_fisico=True)
        if solo_positivos and solo_positivos.lower() == 'true':
            qs = qs.filter(cantidad__gt=0)

        return qs.order_by('articulo__cod_articulo', 'deposito__nombre')


# ─────────────────────────────────────────────
#  VISTAS ADMIN (Django templates — se mantienen)
# ─────────────────────────────────────────────

@staff_member_required
def kardex_articulo_view(request, articulo_id):
    articulo = get_object_or_404(Articulo, pk=articulo_id)
    movimientos = MovimientoStockLedger.objects.filter(
        articulo=articulo
    ).select_related('deposito', 'tipo_stock', 'usuario').order_by('fecha_movimiento', 'pk')

    deposito_id = request.GET.get('deposito')
    if deposito_id:
        movimientos = movimientos.filter(deposito_id=deposito_id)

    filas = []
    saldo_acumulado = 0
    for mov in movimientos:
        impacto = 0
        if mov.tipo_stock and mov.tipo_stock.es_fisico:
            impacto = float(mov.cantidad)
        saldo_acumulado += impacto
        filas.append({
            'fecha': mov.fecha_movimiento,
            'origen': mov.origen_sistema,
            'referencia': mov.origen_referencia,
            'deposito': mov.deposito.nombre if mov.deposito else '',
            'tipo': mov.tipo_stock.nombre if mov.tipo_stock else '',
            'entrada': float(mov.cantidad) if mov.cantidad > 0 else 0,
            'salida': float(abs(mov.cantidad)) if mov.cantidad < 0 else 0,
            'saldo': round(saldo_acumulado, 3),
            'usuario': mov.usuario,
        })

    context = {
        'articulo': articulo,
        'filas': filas,
        'saldo_final': saldo_acumulado,
        'title': f"Ficha de Stock (Kardex): {articulo.descripcion}",
    }
    return render(request, 'admin/inventario/articulo/kardex.html', context)


@staff_member_required
def reporte_valorizacion_view(request):
    deposito_id = request.GET.get('deposito')
    queryset = BalanceStock.objects.filter(
        tipo_stock__codigo='REAL',
        cantidad__gt=0
    ).select_related('articulo', 'deposito', 'articulo__precio_costo_moneda')

    if deposito_id:
        queryset = queryset.filter(deposito_id=deposito_id)
        from .models import Deposito as D
        nombre_deposito = D.objects.get(pk=deposito_id).nombre
    else:
        nombre_deposito = "TODOS LOS DEPÓSITOS"

    lineas = []
    total_valor = 0
    for balance in queryset:
        art = balance.articulo
        subtotal = float(balance.cantidad) * float(art.precio_costo_monto or 0)
        total_valor += subtotal
        lineas.append({
            'codigo': art.cod_articulo,
            'descripcion': art.descripcion,
            'rubro': art.rubro.nombre if art.rubro else '-',
            'deposito': balance.deposito.nombre,
            'cantidad': balance.cantidad,
            'costo': art.precio_costo_monto,
            'total': subtotal,
        })

    context = {
        'lineas': lineas,
        'total_valor': total_valor,
        'depositos': Deposito.objects.all(),
        'deposito_actual': int(deposito_id) if deposito_id else None,
        'nombre_deposito': nombre_deposito,
        'fecha_emision': timezone.now(),
        'moneda': 'ARS',
        'title': "Reporte de Valorización de Inventario",
    }
    return render(request, 'admin/inventario/balance_stock/reporte_valorizacion.html', context)


# ─────────────────────────────────────────────
#  VALORIZACIÓN — API REST (Fase 3)
# ─────────────────────────────────────────────

class ValorizacionView(viewsets.ViewSet):
    """
    GET  /api/inventario/valorizacion/          → datos JSON para la tabla
    GET  /api/inventario/valorizacion/exportar_excel/
    GET  /api/inventario/valorizacion/exportar_pdf/
    """

    def list(self, request):
        deposito_id  = request.query_params.get('deposito')
        rubro_id     = request.query_params.get('rubro')
        solo_positivo = request.query_params.get('solo_positivo', 'true').lower() == 'true'

        qs = BalanceStock.objects.filter(
            tipo_stock__codigo='REAL',
        ).select_related(
            'articulo', 'deposito',
            'articulo__rubro', 'articulo__marca',
            'articulo__precio_costo_moneda', 'articulo__precio_venta_moneda',
        )

        if solo_positivo:
            qs = qs.filter(cantidad__gt=0)
        if deposito_id:
            qs = qs.filter(deposito_id=deposito_id)
        if rubro_id:
            qs = qs.filter(articulo__rubro_id=rubro_id)

        lineas = []
        total_cantidad  = Decimal('0')
        total_costo     = Decimal('0')
        total_venta     = Decimal('0')

        for b in qs:
            art      = b.articulo
            cantidad = Decimal(str(b.cantidad))
            costo_u  = Decimal(str(art.precio_costo_monto or 0))
            venta_u  = Decimal(str(art.precio_venta_monto or 0))
            sub_costo = cantidad * costo_u
            sub_venta = cantidad * venta_u

            total_cantidad += cantidad
            total_costo    += sub_costo
            total_venta    += sub_venta

            lineas.append({
                'articulo_id':   art.pk,
                'cod_articulo':  art.cod_articulo,
                'descripcion':   art.descripcion,
                'rubro':         art.rubro.nombre if art.rubro else '—',
                'marca':         art.marca.nombre if art.marca else '—',
                'deposito':      b.deposito.nombre,
                'cantidad':      str(cantidad),
                'costo_unitario':str(costo_u),
                'venta_unitario':str(venta_u),
                'subtotal_costo':str(sub_costo.quantize(Decimal('0.01'))),
                'subtotal_venta':str(sub_venta.quantize(Decimal('0.01'))),
                'utilidad_pct':  str(art.utilidad or 0),
            })

        return Response({
            'lineas':         lineas,
            'total_cantidad': str(total_cantidad),
            'total_costo':    str(total_costo.quantize(Decimal('0.01'))),
            'total_venta':    str(total_venta.quantize(Decimal('0.01'))),
            'fecha_emision':  timezone.now().isoformat(),
            'depositos':      [{'id': d.id, 'nombre': d.nombre}
                               for d in Deposito.objects.all().order_by('nombre')],
        })

    @action(detail=False, methods=['get'], url_path='exportar_excel')
    def exportar_excel(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, numbers
        from django.http import HttpResponse

        deposito_id = request.query_params.get('deposito')

        qs = BalanceStock.objects.filter(
            tipo_stock__codigo='REAL', cantidad__gt=0
        ).select_related('articulo', 'deposito', 'articulo__rubro', 'articulo__precio_costo_moneda')
        if deposito_id:
            qs = qs.filter(deposito_id=deposito_id)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Valorización'

        # Estilos
        header_font  = Font(bold=True, color='FFFFFF', size=11)
        header_fill  = PatternFill('solid', fgColor='1E3A8A')
        total_fill   = PatternFill('solid', fgColor='DBEAFE')
        total_font   = Font(bold=True, size=11)
        center       = Alignment(horizontal='center')
        right        = Alignment(horizontal='right')
        money_fmt    = '#,##0.00'

        headers = ['Código', 'Descripción', 'Rubro', 'Depósito',
                   'Cantidad', 'Costo Unit.', 'Venta Unit.',
                   'Subtotal Costo', 'Subtotal Venta', 'Utilidad %']
        col_widths = [12, 40, 18, 18, 12, 14, 14, 16, 16, 12]

        for i, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font   = header_font
            cell.fill   = header_fill
            cell.alignment = center
            ws.column_dimensions[cell.column_letter].width = w

        total_costo = Decimal('0')
        total_venta = Decimal('0')
        row = 2

        for b in qs:
            art      = b.articulo
            cantidad = Decimal(str(b.cantidad))
            costo_u  = Decimal(str(art.precio_costo_monto or 0))
            venta_u  = Decimal(str(art.precio_venta_monto or 0))
            sub_c    = float((cantidad * costo_u).quantize(Decimal('0.01')))
            sub_v    = float((cantidad * venta_u).quantize(Decimal('0.01')))
            total_costo += Decimal(str(sub_c))
            total_venta += Decimal(str(sub_v))

            data = [
                art.cod_articulo, art.descripcion,
                art.rubro.nombre if art.rubro else '',
                b.deposito.nombre,
                float(cantidad), float(costo_u), float(venta_u),
                sub_c, sub_v,
                float(art.utilidad or 0),
            ]
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                if col in (5, 6, 7, 8, 9, 10):
                    cell.number_format = money_fmt
                    cell.alignment = right
            row += 1

        # Fila de totales
        ws.cell(row=row, column=4, value='TOTAL').font = total_font
        for col in (5, 8, 9):
            c = ws.cell(row=row, column=col)
            c.font = total_font
            c.fill = total_fill
            c.number_format = money_fmt
        ws.cell(row=row, column=8).value = float(total_costo.quantize(Decimal('0.01')))
        ws.cell(row=row, column=9).value = float(total_venta.quantize(Decimal('0.01')))

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="valorizacion.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=['get'], url_path='exportar_pdf')
    def exportar_pdf(self, request):
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from django.http import HttpResponse
        import io

        deposito_id = request.query_params.get('deposito')

        qs = BalanceStock.objects.filter(
            tipo_stock__codigo='REAL', cantidad__gt=0
        ).select_related('articulo', 'deposito', 'articulo__rubro')
        if deposito_id:
            qs = qs.filter(deposito_id=deposito_id)

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                                leftMargin=1.5*cm, rightMargin=1.5*cm,
                                topMargin=2*cm, bottomMargin=1.5*cm)

        styles  = getSampleStyleSheet()
        title_s = ParagraphStyle('title', parent=styles['Heading1'],
                                 fontSize=14, textColor=colors.HexColor('#1E3A8A'))
        sub_s   = ParagraphStyle('sub', parent=styles['Normal'],
                                 fontSize=9, textColor=colors.HexColor('#64748B'))

        story = [
            Paragraph('Reporte de Valorización de Inventario', title_s),
            Paragraph(f'Generado: {timezone.now().strftime("%d/%m/%Y %H:%M")}', sub_s),
            Spacer(1, 0.4*cm),
        ]

        header = ['Código', 'Descripción', 'Rubro', 'Depósito', 'Cantidad', 'Costo Unit.', 'Subtotal']
        rows   = [header]
        total_val = Decimal('0')

        for b in qs:
            art    = b.articulo
            cant   = Decimal(str(b.cantidad))
            costo  = Decimal(str(art.precio_costo_monto or 0))
            sub    = (cant * costo).quantize(Decimal('0.01'))
            total_val += sub
            rows.append([
                art.cod_articulo,
                art.descripcion[:45],
                art.rubro.nombre[:18] if art.rubro else '—',
                b.deposito.nombre[:18],
                f"{float(cant):,.3f}",
                f"${float(costo):,.2f}",
                f"${float(sub):,.2f}",
            ])

        rows.append(['', '', '', 'TOTAL', '', '', f"${float(total_val):,.2f}"])

        col_widths = [2.5*cm, 8*cm, 4*cm, 4*cm, 2.5*cm, 3*cm, 3.5*cm]
        t = Table(rows, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND',   (0,0), (-1,0),  colors.HexColor('#1E3A8A')),
            ('TEXTCOLOR',    (0,0), (-1,0),  colors.white),
            ('FONTNAME',     (0,0), (-1,0),  'Helvetica-Bold'),
            ('FONTSIZE',     (0,0), (-1,0),  8),
            ('FONTSIZE',     (0,1), (-1,-2), 7),
            ('ROWBACKGROUNDS',(0,1),(-1,-2), [colors.white, colors.HexColor('#F8FAFC')]),
            ('FONTNAME',     (0,-1),(-1,-1), 'Helvetica-Bold'),
            ('BACKGROUND',   (0,-1),(-1,-1), colors.HexColor('#DBEAFE')),
            ('GRID',         (0,0), (-1,-1), 0.3, colors.HexColor('#E2E8F0')),
            ('ALIGN',        (4,0), (-1,-1), 'RIGHT'),
            ('VALIGN',       (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING',   (0,0), (-1,-1), 4),
            ('BOTTOMPADDING',(0,0), (-1,-1), 4),
        ]))
        story.append(t)

        doc.build(story)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="valorizacion.pdf"'
        return response


# ─────────────────────────────────────────────
#  ACTUALIZACIÓN MASIVA DE PRECIOS (Fase 3)
# ─────────────────────────────────────────────

class ActualizacionPreciosView(viewsets.ViewSet):
    """
    POST /api/inventario/actualizar-precios/preview/
         → simula el aumento, devuelve lista de artículos afectados con precio antes/después

    POST /api/inventario/actualizar-precios/aplicar/
         → aplica el aumento definitivamente
    """

    @action(detail=False, methods=['post'], url_path='preview')
    def preview(self, request):
        return self._procesar(request, dry_run=True)

    @action(detail=False, methods=['post'], url_path='aplicar')
    def aplicar(self, request):
        return self._procesar(request, dry_run=False)

    @action(detail=False, methods=['post'], url_path='precios_actuales')
    def precios_actuales(self, request):
        """
        POST /api/inventario/actualizar-precios/precios_actuales/
        Devuelve los precios reales actuales de una lista de artículos,
        sin simular ni aplicar ningún ajuste. Se usa para refrescar la
        tabla después de aplicar una actualización masiva.

        Body: { articulo_ids: [1, 2, 3, ...] }
        """
        articulo_ids = request.data.get('articulo_ids', [])
        if not articulo_ids:
            return Response({'articulos': [], 'total': 0})

        qs = Articulo.objects.filter(pk__in=articulo_ids).select_related(
            'rubro', 'marca', 'precio_costo_moneda', 'precio_venta_moneda'
        )

        articulos = []
        for art in qs:
            costo = Decimal(str(art.precio_costo_monto or 0))
            venta = Decimal(str(art.precio_venta_monto or 0))
            articulos.append({
                'id':              art.pk,
                'cod_articulo':    art.cod_articulo,
                'descripcion':     art.descripcion,
                'rubro':           art.rubro.nombre if art.rubro else '—',
                'costo_antes':     str(costo),
                'costo_nuevo':     str(costo),
                'venta_antes':     str(venta),
                'venta_nuevo':     str(venta),
                'diferencia_venta': '0.00',
            })

        return Response({'articulos': articulos, 'total': len(articulos)})

    def _procesar(self, request, dry_run):
        """
        Parámetros esperados en el body:
          tipo_ajuste: 'porcentaje' | 'monto_fijo' | 'nuevo_costo'
          valor:       número (% o monto según tipo)
          campo:       'costo' | 'venta' | 'ambos'  (qué precio se modifica)
          recalcular_venta: bool — si True y campo='costo', recalcula venta desde utilidad
          filtros:
            rubro_id:    int (opcional)
            marca_id:    int (opcional)
            articulo_ids:[int] (opcional — lista explícita)
            solo_activos: bool (default True)
        """
        from decimal import ROUND_HALF_UP
        from django.db import transaction

        data          = request.data
        tipo_ajuste   = data.get('tipo_ajuste', 'porcentaje')
        valor         = Decimal(str(data.get('valor', 0)))
        campo         = data.get('campo', 'venta')  # 'costo', 'venta', 'ambos'
        recalc_venta  = bool(data.get('recalcular_venta', True))

        filtros       = data.get('filtros', {})
        rubro_id      = filtros.get('rubro_id')
        marca_id      = filtros.get('marca_id')
        articulo_ids  = filtros.get('articulo_ids', [])
        solo_activos  = filtros.get('solo_activos', True)

        if valor == 0:
            return Response({'error': 'El valor del ajuste no puede ser 0.'}, status=400)

        qs = Articulo.objects.select_related('precio_costo_moneda', 'precio_venta_moneda')
        if solo_activos:
            qs = qs.filter(is_active=True)
        if rubro_id:
            qs = qs.filter(rubro_id=rubro_id)
        if marca_id:
            qs = qs.filter(marca_id=marca_id)
        if articulo_ids:
            qs = qs.filter(pk__in=articulo_ids)

        def aplicar_ajuste(precio_actual):
            if tipo_ajuste == 'porcentaje':
                return precio_actual * (1 + valor / Decimal('100'))
            elif tipo_ajuste == 'monto_fijo':
                return precio_actual + valor
            elif tipo_ajuste == 'nuevo_costo':
                # valor ES el nuevo precio absoluto
                return valor
            return precio_actual

        resultado = []
        actualizados = []

        QUANT = Decimal('0.01')

        for art in qs:
            costo_antes = Decimal(str(art.precio_costo_monto or 0))
            venta_antes = Decimal(str(art.precio_venta_monto or 0))

            costo_nuevo = costo_antes
            venta_nuevo = venta_antes

            if campo in ('costo', 'ambos'):
                costo_nuevo = aplicar_ajuste(costo_antes).quantize(QUANT, ROUND_HALF_UP)

            if campo == 'venta':
                venta_nuevo = aplicar_ajuste(venta_antes).quantize(QUANT, ROUND_HALF_UP)
            elif campo == 'ambos':
                venta_nuevo = aplicar_ajuste(venta_antes).quantize(QUANT, ROUND_HALF_UP)
            elif campo == 'costo' and recalc_venta and art.utilidad:
                # Recalcula el precio de venta desde el nuevo costo + utilidad existente
                venta_nuevo = (costo_nuevo * (1 + Decimal(str(art.utilidad)) / 100)).quantize(QUANT, ROUND_HALF_UP)

            resultado.append({
                'id':           art.pk,
                'cod_articulo': art.cod_articulo,
                'descripcion':  art.descripcion,
                'rubro':        art.rubro.nombre if art.rubro else '—',
                'costo_antes':  str(costo_antes),
                'costo_nuevo':  str(costo_nuevo),
                'venta_antes':  str(venta_antes),
                'venta_nuevo':  str(venta_nuevo),
                'diferencia_venta': str((venta_nuevo - venta_antes).quantize(QUANT)),
            })

            if not dry_run:
                actualizados.append((art.pk, costo_nuevo, venta_nuevo))

        if not dry_run:
            with transaction.atomic():
                for art_pk, costo_nuevo, venta_nuevo in actualizados:
                    update_kwargs = {}
                    if campo in ('costo', 'ambos'):
                        update_kwargs['precio_costo_monto'] = costo_nuevo
                    if campo in ('venta', 'ambos') or (campo == 'costo' and recalc_venta):
                        update_kwargs['precio_venta_monto'] = venta_nuevo
                    if update_kwargs:
                        # Usamos queryset.update() en lugar de instance.save() para
                        # bypasear el save() del modelo, que recalcularía precio_venta_monto
                        # automáticamente desde precio_costo_monto * (1 + utilidad/100),
                        # sobreescribiendo el valor que acabamos de calcular.
                        Articulo.objects.filter(pk=art_pk).update(**update_kwargs)

        return Response({
            'dry_run':          dry_run,
            'total_afectados':  len(resultado),
            'tipo_ajuste':      tipo_ajuste,
            'valor':            str(valor),
            'campo':            campo,
            'articulos':        resultado,
        })